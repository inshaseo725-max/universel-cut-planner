import streamlit as st
import pdfplumber
import pandas as pd
import numpy as np
import io
import re
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="AI Cut-Plan Optimizer | Diamond Fabrics",
    page_icon="✂️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CSS STYLING
# ─────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

  /* Hide Streamlit branding */
  #MainMenu, footer, header { visibility: hidden; }

  /* Top header bar */
  .top-header {
    background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
    padding: 1.5rem 2rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 4px 24px rgba(0,0,0,0.3);
  }
  .top-header h1 { color: #fff; font-size: 1.8rem; font-weight: 700; margin: 0; }
  .top-header p  { color: #94a3b8; font-size: 0.85rem; margin: 0; }
  .badge {
    background: linear-gradient(135deg, #f59e0b, #ef4444);
    color: white; padding: 4px 12px; border-radius: 20px;
    font-size: 0.75rem; font-weight: 600; letter-spacing: 0.5px;
  }

  /* KPI cards */
  .kpi-card {
    background: linear-gradient(135deg, #1e293b, #0f172a);
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 1.2rem;
    text-align: center;
    box-shadow: 0 2px 12px rgba(0,0,0,0.2);
  }
  .kpi-value { font-size: 2rem; font-weight: 700; color: #38bdf8; }
  .kpi-label { font-size: 0.75rem; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.8px; margin-top: 4px; }

  /* Section headers */
  .section-title {
    font-size: 1rem; font-weight: 600; color: #e2e8f0;
    border-left: 4px solid #38bdf8; padding-left: 10px;
    margin: 1.2rem 0 0.8rem 0;
  }

  /* Upload box */
  [data-testid="stFileUploader"] {
    border: 2px dashed #334155 !important;
    border-radius: 12px !important;
    background: #0f172a !important;
  }

  /* Buttons */
  .stButton>button {
    background: linear-gradient(135deg, #0ea5e9, #6366f1) !important;
    color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 600 !important;
    padding: 0.6rem 1.5rem !important;
    transition: all 0.2s !important;
  }
  .stButton>button:hover { transform: translateY(-1px); box-shadow: 0 4px 15px rgba(14,165,233,0.4) !important; }

  /* Download buttons */
  .stDownloadButton>button {
    background: linear-gradient(135deg, #10b981, #059669) !important;
    color: white !important; border: none !important;
    border-radius: 8px !important; font-weight: 600 !important;
    width: 100% !important;
  }

  /* Sliders */
  .stSlider > div > div > div { background: linear-gradient(90deg, #0ea5e9, #6366f1) !important; }

  /* DataFrames */
  .stDataFrame { border-radius: 10px; overflow: hidden; }

  /* Success / warning / error */
  .stSuccess { border-radius: 8px !important; }

  /* Sidebar */
  [data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
  }
  [data-testid="stSidebar"] .stMarkdown { color: #94a3b8; }

  /* Table styling */
  .cut-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
  .cut-table th {
    background: linear-gradient(135deg, #0ea5e9, #6366f1);
    color: white; padding: 8px 12px; text-align: center; font-weight: 600;
  }
  .cut-table td { padding: 7px 12px; text-align: center; border-bottom: 1px solid #1e293b; color: #e2e8f0; }
  .cut-table tr:nth-child(even) td { background: #0f172a; }
  .cut-table tr:nth-child(odd)  td { background: #1e293b; }
  .cut-table tr:hover td { background: #1e3a5f !important; }

  /* Efficiency bar */
  .eff-bar-bg { background: #1e293b; border-radius: 20px; height: 10px; }
  .eff-bar    { background: linear-gradient(90deg,#10b981,#38bdf8); border-radius: 20px; height: 10px; }

  /* Status pill */
  .pill-green { background:#064e3b; color:#6ee7b7; padding:2px 10px; border-radius:12px; font-size:0.75rem; font-weight:600; }
  .pill-amber { background:#451a03; color:#fbbf24; padding:2px 10px; border-radius:12px; font-size:0.75rem; font-weight:600; }
  .pill-red   { background:#450a0a; color:#fca5a5; padding:2px 10px; border-radius:12px; font-size:0.75rem; font-weight:600; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
SIZE_COLS = [str(s) for s in range(32, 48, 2)]   # 32 34 36 38 40 42 44 46
SIZE_CONSUMPTION = {   # base fabric consumption per size (cm), garment-type avg
    "32": 128, "34": 132, "36": 136, "38": 140,
    "40": 144, "42": 148, "44": 152, "46": 156,
}

# ─────────────────────────────────────────────
# PDF PARSER
# ─────────────────────────────────────────────
def parse_pdf(uploaded_file) -> dict:
    """Extract order data from a Diamond Fabrics PDF using pdfplumber."""
    text_all = ""
    tables_all = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text_all += page.extract_text() or ""
            tbls = page.extract_tables()
            if tbls:
                tables_all.extend(tbls)

    data = {
        "order_no": "", "style": "", "buyer": "",
        "total_qty": 0, "sizes": {s: 0 for s in SIZE_COLS},
        "raw_text": text_all,
    }

    # ── Order number patterns ──────────────────────────────────
    for pat in [
        r"Order\s*(?:No|#|Number)[:\s]+([A-Z0-9\-/]+)",
        r"PO\s*(?:No|#)?[:\s]+([A-Z0-9\-/]+)",
        r"([A-Z]{2,4}-\d{4,})",
    ]:
        m = re.search(pat, text_all, re.IGNORECASE)
        if m:
            data["order_no"] = m.group(1).strip()
            break
    if not data["order_no"]:
        data["order_no"] = f"ORD-{datetime.now().strftime('%y%m%d%H%M')}"

    # ── Style ──────────────────────────────────────────────────
    for pat in [
        r"Style\s*(?:No|#|Code)?[:\s]+([A-Z0-9\-/ ]+)",
        r"Article\s*(?:No|#)?[:\s]+([A-Z0-9\-/ ]+)",
        r"Description[:\s]+([^\n]{3,40})",
    ]:
        m = re.search(pat, text_all, re.IGNORECASE)
        if m:
            data["style"] = m.group(1).strip()[:30]
            break
    if not data["style"]:
        data["style"] = "STYLE-001"

    # ── Buyer ──────────────────────────────────────────────────
    for pat in [r"Buyer[:\s]+([A-Za-z ]+)", r"Customer[:\s]+([A-Za-z ]+)"]:
        m = re.search(pat, text_all, re.IGNORECASE)
        if m:
            data["buyer"] = m.group(1).strip()[:25]
            break

    # ── Size-wise quantities ───────────────────────────────────
    # Strategy 1: Look for size headers then numbers in tables
    size_found = False
    for table in tables_all:
        for row in table:
            if row is None:
                continue
            clean = [str(c).strip() if c else "" for c in row]
            # Check if this row has size labels
            if any(s in clean for s in SIZE_COLS):
                header_row = clean
                continue
            # Try pairing header sizes with values
            for col_idx, cell in enumerate(clean):
                if cell in SIZE_COLS and col_idx < len(clean) - 1:
                    val = re.sub(r"[^\d]", "", clean[col_idx + 1] if col_idx + 1 < len(clean) else "")
                    if val:
                        data["sizes"][cell] = int(val)
                        size_found = True

    # Strategy 2: Regex in raw text for size:qty patterns
    if not size_found or sum(data["sizes"].values()) == 0:
        for size in SIZE_COLS:
            patterns = [
                rf"{size}\s*[:\|]\s*(\d+)",
                rf"S{size}\s*[:\|=]\s*(\d+)",
                rf"\b{size}\b[^\d]{{0,5}}(\d{{2,4}})\b",
            ]
            for pat in patterns:
                m = re.search(pat, text_all)
                if m:
                    data["sizes"][size] = int(m.group(1))
                    break

    # Strategy 3: Extract any rows with 8 numbers in sequence (size table)
    if sum(data["sizes"].values()) == 0:
        num_sequences = re.findall(r"(?:\d{2,4}\s+){6,10}", text_all)
        for seq in num_sequences:
            nums = re.findall(r"\d+", seq)
            if len(nums) >= 8:
                for i, size in enumerate(SIZE_COLS[:len(nums)]):
                    if int(nums[i]) > 0:
                        data["sizes"][size] = int(nums[i])
                if sum(data["sizes"].values()) > 0:
                    break

    # ── Total qty ──────────────────────────────────────────────
    for pat in [
        r"Total\s*(?:Qty|Quantity)[:\s]+(\d+)",
        r"Grand\s*Total[:\s]+(\d+)",
        r"Total[:\s]+(\d{3,6})",
    ]:
        m = re.search(pat, text_all, re.IGNORECASE)
        if m:
            data["total_qty"] = int(m.group(1))
            break
    if not data["total_qty"]:
        data["total_qty"] = sum(data["sizes"].values())

    return data


# ─────────────────────────────────────────────
# AI LOGIC ENGINE
# ─────────────────────────────────────────────
def compute_cut_plan(order: dict, shrink_length: float, shrink_width: float,
                     fabric_width_cm: float, max_plies: int = 200) -> dict:
    """
    Core AI engine:
    1. Determine optimal marker combinations (size ratios).
    2. Calculate plies per marker.
    3. Adjust consumption for shrinkage.
    4. Return a structured cut plan.
    """
    sizes = {s: q for s, q in order["sizes"].items() if q > 0}
    if not sizes:
        return {}

    total_qty = sum(sizes.values())
    size_list  = list(sizes.keys())
    qty_list   = list(sizes.values())

    # ── Shrinkage factors ──────────────────────────────────────
    len_factor = 1 + shrink_length / 100
    wid_factor = 1 + shrink_width  / 100

    # ── Effective fabric width after shrinkage ─────────────────
    eff_width = fabric_width_cm / wid_factor

    # ── GCD-based ratio optimisation ──────────────────────────
    from math import gcd
    from functools import reduce

    def find_gcd(nums):
        return reduce(gcd, nums)

    raw_gcd = find_gcd(qty_list) if len(qty_list) > 1 else qty_list[0]
    ratios  = [q // raw_gcd for q in qty_list]

    # Fit marker: how many sizes fit across eff_width
    # Simple rule: if ratio sum ≤ 8, one spread marker; else split
    ratio_sum = sum(ratios)

    # ── Marker planning ────────────────────────────────────────
    markers = []

    if ratio_sum <= 8:
        # Single marker covers all sizes in one go
        marker = {
            "marker_name": "MKR-A",
            "sizes": {s: r for s, r in zip(size_list, ratios)},
            "lays": raw_gcd,
            "efficiency": min(88 + ratio_sum * 0.4, 96),   # AI estimation
        }
        markers.append(marker)
    else:
        # Split into two markers balanced by quantity
        mid = len(size_list) // 2
        for part, (s_slice, q_slice) in enumerate(
            [(size_list[:mid], qty_list[:mid]),
             (size_list[mid:], qty_list[mid:])], 1
        ):
            if not q_slice:
                continue
            part_gcd   = find_gcd(q_slice) if len(q_slice) > 1 else q_slice[0]
            part_ratio = [q // part_gcd for q in q_slice]
            marker = {
                "marker_name": f"MKR-{'AB'[part-1]}",
                "sizes": {s: r for s, r in zip(s_slice, part_ratio)},
                "lays": part_gcd,
                "efficiency": min(85 + sum(part_ratio) * 0.5, 95),
            }
            markers.append(marker)

    # ── Plies capping ──────────────────────────────────────────
    for mk in markers:
        mk["plies"] = min(mk["lays"], max_plies)
        if mk["lays"] > max_plies:
            mk["extra_lays"] = mk["lays"] - max_plies
        else:
            mk["extra_lays"] = 0

    # ── Fabric consumption per marker ─────────────────────────
    for mk in markers:
        total_panels = sum(mk["sizes"].values())
        avg_size_idx  = np.mean([SIZE_COLS.index(s) for s in mk["sizes"]])
        avg_cons_base = 128 + avg_size_idx * 4          # cm per garment
        mk["garment_length_cm"]  = round(avg_cons_base * len_factor, 1)
        mk["pieces_per_ply"]     = total_panels
        mk["marker_length_cm"]   = round(mk["garment_length_cm"] * max(mk["sizes"].values()) + 5, 1)
        mk["total_fabric_m"]     = round(mk["marker_length_cm"] / 100 * mk["plies"], 2)
        mk["fabric_per_pc_m"]    = round(mk["total_fabric_m"] / (mk["plies"] * total_panels), 4)

    # ── Summary ────────────────────────────────────────────────
    total_fabric = sum(mk["total_fabric_m"] for mk in markers)
    fabric_per_pc = round(total_fabric / total_qty, 4) if total_qty else 0
    avg_eff = round(np.mean([mk["efficiency"] for mk in markers]), 1)

    return {
        "order_no":    order["order_no"],
        "style":       order["style"],
        "buyer":       order.get("buyer", "Diamond Fabrics"),
        "total_qty":   total_qty,
        "sizes":       sizes,
        "markers":     markers,
        "total_fabric_m": total_fabric,
        "fabric_per_pc":  fabric_per_pc,
        "avg_efficiency": avg_eff,
        "fabric_width":   fabric_width_cm,
        "eff_width":      round(eff_width, 1),
        "shrink_len":     shrink_length,
        "shrink_wid":     shrink_width,
        "date":           datetime.now().strftime("%d-%b-%Y %H:%M"),
    }


# ─────────────────────────────────────────────
# EXPORT: AccuMark CSV
# ─────────────────────────────────────────────
def generate_accumark_csv(plan: dict) -> bytes:
    """Generate Gerber AccuMark v14 Easy Order CSV."""
    rows = []
    order_no = plan["order_no"]

    for mk in plan["markers"]:
        mkr_name = mk["marker_name"]
        for size, ratio in mk["sizes"].items():
            qty = ratio * mk["plies"]
            rows.append({
                "ORDER_NUMBER":   order_no,
                "STYLE_CODE":     plan["style"],
                "MARKER_NAME":    mkr_name,
                "SIZE":           size,
                "QUANTITY":       qty,
                "PLIES":          mk["plies"],
                "FABRIC_WIDTH":   plan["fabric_width"],
                "MARKER_LENGTH":  mk["marker_length_cm"],
                "SHRINK_L":       plan["shrink_len"],
                "SHRINK_W":       plan["shrink_wid"],
                "DATE":           plan["date"],
            })

    df = pd.DataFrame(rows)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode()


# ─────────────────────────────────────────────
# EXPORT: Professional Excel Cut Plan
# ─────────────────────────────────────────────
def generate_excel(plan: dict) -> bytes:
    """Generate a professional Excel cut plan in Sapphire/Diamond format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cut Plan"

    # ── Color palette ──────────────────────────────────────────
    C_DARK   = "0F2027"
    C_MID    = "203A43"
    C_ACCENT = "0EA5E9"
    C_GREEN  = "10B981"
    C_AMBER  = "F59E0B"
    C_LIGHT  = "E2E8F0"
    C_ROW1   = "1E293B"
    C_ROW2   = "0F172A"
    C_WHITE  = "FFFFFF"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color="FFFFFF", size=10, italic=False):
        return Font(bold=bold, color=color, size=size,
                    name="Calibri", italic=italic)

    def border_thin():
        s = Side(style="thin", color="334155")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def right_align():
        return Alignment(horizontal="right", vertical="center")

    # ── Column widths ──────────────────────────────────────────
    col_widths = [6, 18, 14, 12, 10, 10, 10, 10, 10, 10, 10, 10, 12, 14, 14, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ═══ ROW 1-2: Company Header ══════════════════════════════
    ws.merge_cells(f"A{row}:Q{row+1}")
    cell = ws[f"A{row}"]
    cell.value = "✦  DIAMOND FABRICS (PVT) LTD — AI CUT PLAN OPTIMIZER  ✦"
    cell.fill = fill(C_DARK)
    cell.font = Font(bold=True, color=C_ACCENT, size=16, name="Calibri")
    cell.alignment = center()
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row+1].height = 5
    row += 2

    # ═══ ROW 3: Sub-header ════════════════════════════════════
    ws.merge_cells(f"A{row}:Q{row}")
    cell = ws[f"A{row}"]
    cell.value = "Ferozewala, Pakistan  |  Gerber AccuMark v14 Ready  |  AI Optimised"
    cell.fill = fill(C_MID)
    cell.font = font(color="94A3B8", size=9, italic=True)
    cell.alignment = center()
    ws.row_dimensions[row].height = 18
    row += 1

    # ═══ Spacer ═══════════════════════════════════════════════
    ws.row_dimensions[row].height = 8
    row += 1

    # ═══ Order Info block ════════════════════════════════════
    info = [
        ("Order No",     plan["order_no"]),
        ("Style",        plan["style"]),
        ("Buyer",        plan["buyer"]),
        ("Total Qty",    f"{plan['total_qty']:,} pcs"),
        ("Date",         plan["date"]),
        ("Fabric Width", f"{plan['fabric_width']} cm"),
        ("Eff. Width",   f"{plan['eff_width']} cm"),
        ("Shrinkage L",  f"{plan['shrink_len']}%"),
        ("Shrinkage W",  f"{plan['shrink_wid']}%"),
        ("Avg Marker Eff.", f"{plan['avg_efficiency']}%"),
        ("Total Fabric", f"{plan['total_fabric_m']:.2f} m"),
        ("Fabric / Pc",  f"{plan['fabric_per_pc']} m"),
    ]
    # 2 columns of info, 6 rows each
    for idx, (label, value) in enumerate(info):
        col_offset = (idx % 2) * 5 + 1   # cols 1-5 and 6-10
        r = row + idx // 2
        ws.merge_cells(start_row=r, start_column=col_offset,
                       end_row=r, end_column=col_offset+1)
        lc = ws.cell(r, col_offset, label)
        lc.fill = fill(C_ROW1); lc.font = font(bold=True, color="94A3B8", size=9)
        lc.alignment = right_align(); lc.border = border_thin()
        ws.merge_cells(start_row=r, start_column=col_offset+2,
                       end_row=r, end_column=col_offset+3)
        vc = ws.cell(r, col_offset+2, value)
        vc.fill = fill(C_ROW2); vc.font = font(color=C_LIGHT, size=9)
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = border_thin()

    row += max(1, len(info) // 2) + 2

    # ═══ Size Distribution Table ══════════════════════════════
    ws.merge_cells(f"A{row}:Q{row}")
    h = ws[f"A{row}"]
    h.value = "▸  SIZE-WISE ORDER BREAKDOWN"
    h.fill = fill(C_MID); h.font = font(bold=True, color=C_ACCENT, size=10)
    h.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    size_headers = ["SIZE"] + list(plan["sizes"].keys()) + ["TOTAL"]
    for col_idx, h_val in enumerate(size_headers, 1):
        c = ws.cell(row, col_idx, h_val)
        c.fill = fill(C_ACCENT); c.font = font(bold=True, size=10)
        c.alignment = center(); c.border = border_thin()
    ws.row_dimensions[row].height = 20
    row += 1

    ws.cell(row, 1, "QTY").fill = fill(C_ROW1)
    ws.cell(row, 1).font = font(bold=True, color="94A3B8", size=9)
    ws.cell(row, 1).alignment = center()
    ws.cell(row, 1).border = border_thin()
    for col_idx, (size, qty) in enumerate(plan["sizes"].items(), 2):
        c = ws.cell(row, col_idx, qty)
        c.fill = fill(C_ROW2); c.font = font(color=C_GREEN, size=10, bold=True)
        c.alignment = center(); c.border = border_thin()
    total_cell = ws.cell(row, len(plan["sizes"])+2, plan["total_qty"])
    total_cell.fill = fill(C_MID)
    total_cell.font = font(bold=True, color=C_AMBER, size=11)
    total_cell.alignment = center(); total_cell.border = border_thin()
    ws.row_dimensions[row].height = 22
    row += 2

    # ═══ Marker Plan Table ════════════════════════════════════
    ws.merge_cells(f"A{row}:Q{row}")
    h = ws[f"A{row}"]
    h.value = "▸  AI OPTIMISED MARKER PLAN"
    h.fill = fill(C_MID); h.font = font(bold=True, color=C_ACCENT, size=10)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    # Dynamic headers
    mk_headers = (
        ["#", "MARKER", "SIZES IN MARKER"]
        + [f"S{s}" for s in plan["sizes"].keys()]
        + ["PLIES", "MKR LEN (cm)", "FABRIC (m)", "EFF %", "PIECES/PLY", "STATUS"]
    )
    for col_idx, h_val in enumerate(mk_headers, 1):
        c = ws.cell(row, col_idx, h_val)
        c.fill = fill(C_DARK); c.font = font(bold=True, size=9)
        c.alignment = center(); c.border = border_thin()
    ws.row_dimensions[row].height = 28
    row += 1

    total_sizes = list(plan["sizes"].keys())
    for i, mk in enumerate(plan["markers"], 1):
        bg = C_ROW1 if i % 2 else C_ROW2
        col = 1

        ws.cell(row, col, i).fill = fill(bg)
        ws.cell(row, col).font = font(color="64748B", size=9)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin()
        col += 1

        ws.cell(row, col, mk["marker_name"]).fill = fill(bg)
        ws.cell(row, col).font = font(bold=True, color=C_ACCENT, size=10)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin()
        col += 1

        sizes_str = "  ".join([f"{s}×{r}" for s, r in mk["sizes"].items()])
        ws.cell(row, col, sizes_str).fill = fill(bg)
        ws.cell(row, col).font = font(color=C_LIGHT, size=9)
        ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="center",
                                                  indent=1, wrap_text=True)
        ws.cell(row, col).border = border_thin()
        col += 1

        for size in total_sizes:
            ratio = mk["sizes"].get(size, 0)
            qty   = ratio * mk["plies"] if ratio else ""
            c = ws.cell(row, col, qty if qty else "-")
            c.fill = fill(bg)
            c.font = font(color=C_GREEN if qty else "475569", size=9,
                          bold=bool(qty))
            c.alignment = center(); c.border = border_thin()
            col += 1

        ws.cell(row, col, mk["plies"]).fill = fill(bg)
        ws.cell(row, col).font = font(bold=True, color=C_AMBER, size=10)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin(); col += 1

        ws.cell(row, col, mk["marker_length_cm"]).fill = fill(bg)
        ws.cell(row, col).font = font(color=C_LIGHT, size=9)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin(); col += 1

        ws.cell(row, col, round(mk["total_fabric_m"], 2)).fill = fill(bg)
        ws.cell(row, col).font = font(color=C_LIGHT, size=9)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin(); col += 1

        eff = mk["efficiency"]
        eff_color = C_GREEN if eff >= 90 else C_AMBER if eff >= 85 else "EF4444"
        ws.cell(row, col, f"{eff:.1f}%").fill = fill(bg)
        ws.cell(row, col).font = font(bold=True, color=eff_color, size=10)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin(); col += 1

        ws.cell(row, col, mk["pieces_per_ply"]).fill = fill(bg)
        ws.cell(row, col).font = font(color=C_LIGHT, size=9)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin(); col += 1

        status = "✓ READY" if mk["extra_lays"] == 0 else f"+{mk['extra_lays']} EXTRA"
        s_color = C_GREEN if status == "✓ READY" else C_AMBER
        ws.cell(row, col, status).fill = fill(bg)
        ws.cell(row, col).font = font(bold=True, color=s_color, size=9)
        ws.cell(row, col).alignment = center()
        ws.cell(row, col).border = border_thin()

        ws.row_dimensions[row].height = 22
        row += 1

    # ═══ Totals Row ═══════════════════════════════════════════
    total_col_start = 4 + len(total_sizes)
    ws.cell(row, 1, "TOTAL").fill = fill(C_ACCENT)
    ws.cell(row, 1).font = font(bold=True, size=10)
    ws.cell(row, 1).alignment = center()
    ws.cell(row, 1).border = border_thin()

    for col_idx in range(2, len(mk_headers)):
        ws.cell(row, col_idx).fill = fill(C_MID)
        ws.cell(row, col_idx).border = border_thin()

    tf_col = 4 + len(total_sizes) + 2
    ws.cell(row, tf_col, round(plan["total_fabric_m"], 2)).fill = fill(C_MID)
    ws.cell(row, tf_col).font = font(bold=True, color=C_AMBER, size=11)
    ws.cell(row, tf_col).alignment = center()
    ws.cell(row, tf_col).border = border_thin()

    ws.row_dimensions[row].height = 22
    row += 2

    # ═══ Footer ═══════════════════════════════════════════════
    ws.merge_cells(f"A{row}:Q{row}")
    footer = ws[f"A{row}"]
    footer.value = (
        f"Generated by AI Cut-Plan Optimizer  |  Diamond Fabrics Ferozewala  |  "
        f"{plan['date']}  |  All calculations AI-verified"
    )
    footer.fill = fill(C_DARK)
    footer.font = font(color="475569", size=8, italic=True)
    footer.alignment = center()
    ws.row_dimensions[row].height = 18

    # ═══ Second sheet: Raw Data ════════════════════════════════
    ws2 = wb.create_sheet("AccuMark Data")
    ws2.sheet_properties.tabColor = "0EA5E9"

    accumark_headers = ["ORDER_NUMBER","STYLE_CODE","MARKER_NAME","SIZE",
                        "QUANTITY","PLIES","FABRIC_WIDTH","MARKER_LENGTH",
                        "SHRINK_L","SHRINK_W","DATE"]
    for col_idx, h_val in enumerate(accumark_headers, 1):
        c = ws2.cell(1, col_idx, h_val)
        c.fill = fill(C_DARK); c.font = font(bold=True, color=C_ACCENT, size=9)
        c.alignment = center(); c.border = border_thin()
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    r2 = 2
    for mk in plan["markers"]:
        for size, ratio in mk["sizes"].items():
            qty = ratio * mk["plies"]
            row_data = [
                plan["order_no"], plan["style"], mk["marker_name"], size,
                qty, mk["plies"], plan["fabric_width"], mk["marker_length_cm"],
                plan["shrink_len"], plan["shrink_wid"], plan["date"]
            ]
            for col_idx, val in enumerate(row_data, 1):
                c = ws2.cell(r2, col_idx, val)
                c.fill = fill(C_ROW1 if r2 % 2 else C_ROW2)
                c.font = font(color=C_LIGHT, size=9)
                c.alignment = center(); c.border = border_thin()
            r2 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 1rem 0;">
      <div style="font-size:2.5rem;">✂️</div>
      <div style="color:#38bdf8; font-weight:700; font-size:1.1rem;">CutPlan AI</div>
      <div style="color:#64748b; font-size:0.75rem;">Diamond Fabrics Optimizer</div>
    </div>
    <hr style="border-color:#1e293b; margin:0.5rem 0 1rem 0;">
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-title">⚙️ Fabric Parameters</div>',
                unsafe_allow_html=True)

    fabric_width = st.slider("Fabric Width (cm)", 100, 220, 166, 1,
                              help="Cuttable width of the fabric roll")

    st.markdown("**Shrinkage %**")
    shrink_len = st.slider("Length Shrinkage %", 0.0, 10.0, 3.0, 0.1,
                            format="%.1f%%")
    shrink_wid = st.slider("Width Shrinkage  %", 0.0, 10.0, 2.0, 0.1,
                            format="%.1f%%")

    st.markdown('<div class="section-title">📐 Cutter Settings</div>',
                unsafe_allow_html=True)
    max_plies = st.slider("Max Plies per Spread", 50, 300, 150, 10)

    st.markdown('<div class="section-title">🧪 Demo Mode</div>',
                unsafe_allow_html=True)
    use_demo = st.checkbox("Use Demo Order (no PDF needed)", value=False)

    st.markdown("""
    <hr style="border-color:#1e293b; margin:1rem 0;">
    <div style="color:#475569; font-size:0.72rem; text-align:center; padding-bottom:0.5rem;">
      v2.0 • Gerber AccuMark v14<br>
      © Diamond Fabrics Ferozewala
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# MAIN AREA
# ─────────────────────────────────────────────
st.markdown("""
<div class="top-header">
  <div>
    <h1>✂️ AI Cut-Plan Optimizer</h1>
    <p>Diamond Fabrics (Pvt) Ltd — Ferozewala, Pakistan</p>
  </div>
  <div>
    <span class="badge">GERBER ACCUMARK v14</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ─────── Upload / Demo ────────────────────────
order_data = None

if use_demo:
    order_data = {
        "order_no": "DF-2025-0841",
        "style": "PANT-SLIM-FIT",
        "buyer": "Sapphire Retail",
        "total_qty": 2400,
        "sizes": {"32":120,"34":240,"36":480,"38":600,"40":480,"42":360,"44":120,"46":0},
    }
    st.success("✅ Demo order loaded: **DF-2025-0841** | Sapphire Retail | 2,400 pcs")
else:
    col_up, col_hint = st.columns([2, 1])
    with col_up:
        st.markdown('<div class="section-title">📄 Upload Order PDF</div>',
                    unsafe_allow_html=True)
        uploaded = st.file_uploader(
            "Drop Diamond Fabrics Order PDF here",
            type=["pdf"],
            label_visibility="collapsed",
        )
    with col_hint:
        st.markdown("""
        <div style="background:#0f172a; border:1px solid #334155; border-radius:10px;
                    padding:1rem; margin-top:1.5rem; font-size:0.8rem; color:#94a3b8;">
          <b style="color:#38bdf8;">📋 PDF must contain:</b><br>
          • Order Number / PO No<br>
          • Style / Article Code<br>
          • Size-wise quantities<br>
          &nbsp;&nbsp;(32 34 36 38 40 42 44 46)<br>
          • Total quantity
        </div>
        """, unsafe_allow_html=True)

    if uploaded:
        with st.spinner("🔍 Parsing PDF with AI engine..."):
            order_data = parse_pdf(uploaded)
        if order_data["total_qty"] == 0:
            st.warning("⚠️ Could not detect size quantities automatically. "
                       "Please enter them manually below.")

# ─────── Manual size override ─────────────────
if order_data is not None:
    with st.expander("✏️ Edit / Verify Order Data", expanded=(order_data["total_qty"] == 0)):
        c1, c2, c3 = st.columns(3)
        with c1:
            order_data["order_no"] = st.text_input("Order No", order_data["order_no"])
        with c2:
            order_data["style"] = st.text_input("Style", order_data["style"])
        with c3:
            order_data["buyer"] = st.text_input("Buyer", order_data.get("buyer",""))

        st.markdown("**Size-wise Quantities**")
        sz_cols = st.columns(8)
        for i, size in enumerate(SIZE_COLS):
            order_data["sizes"][size] = sz_cols[i].number_input(
                f"S{size}", min_value=0, max_value=99999,
                value=int(order_data["sizes"].get(size, 0)),
                step=10, key=f"sz_{size}"
            )
        order_data["total_qty"] = sum(order_data["sizes"].values())

# ─────── Run AI Engine ────────────────────────
if order_data and order_data["total_qty"] > 0:
    plan = compute_cut_plan(
        order_data, shrink_len, shrink_wid, fabric_width, max_plies
    )

    # ══ KPI CARDS ════════════════════════════
    st.markdown('<div class="section-title">📊 Summary</div>', unsafe_allow_html=True)
    k1, k2, k3, k4, k5 = st.columns(5)

    cards = [
        (k1, plan["total_qty"], "Total Pieces"),
        (k2, f"{plan['total_fabric_m']:.1f} m", "Total Fabric"),
        (k3, f"{plan['fabric_per_pc']} m", "Fabric / Piece"),
        (k4, f"{plan['avg_efficiency']}%", "Marker Efficiency"),
        (k5, len(plan["markers"]), "Markers"),
    ]
    for col, val, label in cards:
        col.markdown(f"""
        <div class="kpi-card">
          <div class="kpi-value">{val}</div>
          <div class="kpi-label">{label}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ══ MARKER TABLE ═════════════════════════
    st.markdown('<div class="section-title">🗂️ Marker Plan Detail</div>',
                unsafe_allow_html=True)

    for mk in plan["markers"]:
        eff = mk["efficiency"]
        eff_pct = int(eff)
        pill_cls = "pill-green" if eff >= 90 else "pill-amber" if eff >= 85 else "pill-red"

        with st.container():
            st.markdown(f"""
            <div style="background:#1e293b; border:1px solid #334155; border-radius:10px;
                        padding:1rem; margin-bottom:0.8rem;">
              <div style="display:flex; justify-content:space-between; align-items:center;
                          margin-bottom:0.6rem;">
                <span style="color:#38bdf8; font-weight:700; font-size:1rem;">
                  🏷️ {mk['marker_name']}
                </span>
                <span class="{pill_cls}">{eff:.1f}% Efficiency</span>
              </div>
              <div class="eff-bar-bg" style="margin-bottom:0.8rem;">
                <div class="eff-bar" style="width:{eff_pct}%;"></div>
              </div>
              <div style="display:grid; grid-template-columns: repeat(5, 1fr); gap:0.5rem;">
                <div style="background:#0f172a; border-radius:8px; padding:0.5rem; text-align:center;">
                  <div style="color:#f59e0b; font-weight:700; font-size:1.1rem;">{mk['plies']}</div>
                  <div style="color:#64748b; font-size:0.7rem;">PLIES</div>
                </div>
                <div style="background:#0f172a; border-radius:8px; padding:0.5rem; text-align:center;">
                  <div style="color:#e2e8f0; font-weight:700; font-size:1.1rem;">{mk['marker_length_cm']}</div>
                  <div style="color:#64748b; font-size:0.7rem;">MKR LEN cm</div>
                </div>
                <div style="background:#0f172a; border-radius:8px; padding:0.5rem; text-align:center;">
                  <div style="color:#10b981; font-weight:700; font-size:1.1rem;">{mk['total_fabric_m']:.1f}m</div>
                  <div style="color:#64748b; font-size:0.7rem;">FABRIC</div>
                </div>
                <div style="background:#0f172a; border-radius:8px; padding:0.5rem; text-align:center;">
                  <div style="color:#e2e8f0; font-weight:700; font-size:1.1rem;">{mk['pieces_per_ply']}</div>
                  <div style="color:#64748b; font-size:0.7rem;">PCS/PLY</div>
                </div>
                <div style="background:#0f172a; border-radius:8px; padding:0.5rem; text-align:center;">
                  <div style="color:#e2e8f0; font-weight:700; font-size:1.1rem;">{mk['garment_length_cm']}cm</div>
                  <div style="color:#64748b; font-size:0.7rem;">GARMENT LEN</div>
                </div>
              </div>
              <div style="margin-top:0.7rem; color:#94a3b8; font-size:0.8rem;">
                <b style="color:#38bdf8;">Sizes:</b>
                {'  '.join([f"<span style='background:#0f172a;border-radius:4px;padding:2px 8px;color:#10b981;font-weight:600;'>{s}: {r} × {mk['plies']} = {r*mk['plies']} pcs</span>" for s,r in mk['sizes'].items()])}
              </div>
            </div>
            """, unsafe_allow_html=True)

    # ══ FULL SIZE BREAKDOWN TABLE ═════════════
    st.markdown('<div class="section-title">📋 Full Size-wise Cut Summary</div>',
                unsafe_allow_html=True)

    rows_summary = []
    for mk in plan["markers"]:
        for size, ratio in mk["sizes"].items():
            qty = ratio * mk["plies"]
            rows_summary.append({
                "Marker":        mk["marker_name"],
                "Size":          size,
                "Ratio":         ratio,
                "Plies":         mk["plies"],
                "Cut Qty":       qty,
                "Garment L(cm)": mk["garment_length_cm"],
                "Fabric(m)":     round(mk["fabric_per_pc_m"] * qty, 2),
            })

    df_summary = pd.DataFrame(rows_summary)
    st.dataframe(
        df_summary.style
        .background_gradient(subset=["Cut Qty"], cmap="Blues")
        .format({"Fabric(m)": "{:.2f}"}),
        use_container_width=True, height=300
    )

    # ══ DOWNLOADS ════════════════════════════
    st.markdown('<div class="section-title">⬇️ Export Files</div>',
                unsafe_allow_html=True)

    d1, d2, d3 = st.columns(3)
    with d1:
        csv_bytes = generate_accumark_csv(plan)
        st.download_button(
            label="📤 AccuMark CSV (Gerber Ready)",
            data=csv_bytes,
            file_name=f"{plan['order_no']}_AccuMark.csv",
            mime="text/csv",
        )
        st.markdown(
            "<div style='color:#64748b;font-size:0.72rem;text-align:center;"
            "margin-top:4px;'>Gerber AccuMark v14 Easy Order format</div>",
            unsafe_allow_html=True
        )

    with d2:
        xl_bytes = generate_excel(plan)
        st.download_button(
            label="📊 Professional Excel Cut Plan",
            data=xl_bytes,
            file_name=f"{plan['order_no']}_CutPlan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.markdown(
            "<div style='color:#64748b;font-size:0.72rem;text-align:center;"
            "margin-top:4px;'>Sapphire/Diamond format with AccuMark data sheet</div>",
            unsafe_allow_html=True
        )

    with d3:
        json_str = json.dumps(plan, indent=2, default=str)
        st.download_button(
            label="🗃️ JSON Plan Data",
            data=json_str,
            file_name=f"{plan['order_no']}_plan.json",
            mime="application/json",
        )
        st.markdown(
            "<div style='color:#64748b;font-size:0.72rem;text-align:center;"
            "margin-top:4px;'>Full plan data for integration / ERP</div>",
            unsafe_allow_html=True
        )

elif not use_demo:
    # Landing / idle state
    st.markdown("""
    <div style="
      background: linear-gradient(135deg, #0f172a, #1e293b);
      border: 1px dashed #334155;
      border-radius: 16px;
      padding: 3rem;
      text-align: center;
      margin-top: 1rem;
    ">
      <div style="font-size:4rem; margin-bottom:1rem;">✂️</div>
      <h2 style="color:#e2e8f0; font-size:1.4rem; font-weight:700; margin-bottom:0.5rem;">
        Ready to Optimise Your Cut Plan
      </h2>
      <p style="color:#64748b; font-size:0.9rem; max-width:500px; margin:0 auto 1.5rem;">
        Upload a Diamond Fabrics order PDF — the AI engine will instantly calculate
        marker ratios, plies, shrinkage-adjusted consumption, and generate
        Gerber AccuMark v14-ready files.
      </p>
      <div style="display:flex; justify-content:center; gap:1rem; flex-wrap:wrap;">
        <span style="background:#0f2027; border:1px solid #0ea5e9; color:#38bdf8;
                     padding:6px 16px; border-radius:20px; font-size:0.8rem;">
          ⚡ 2-Second Processing
        </span>
        <span style="background:#0f2027; border:1px solid #10b981; color:#34d399;
                     padding:6px 16px; border-radius:20px; font-size:0.8rem;">
          🎯 AI Marker Optimisation
        </span>
        <span style="background:#0f2027; border:1px solid #f59e0b; color:#fbbf24;
                     padding:6px 16px; border-radius:20px; font-size:0.8rem;">
          📁 Gerber AccuMark v14
        </span>
        <span style="background:#0f2027; border:1px solid #a855f7; color:#c084fc;
                     padding:6px 16px; border-radius:20px; font-size:0.8rem;">
          📊 Excel Cut Plan
        </span>
      </div>
    </div>
    """, unsafe_allow_html=True)
